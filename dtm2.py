import os
import math
import openpyxl
import pandas as pd 
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText
from tkcalendar import Calendar, DateEntry
from datetime import date

root = Tk()

root.title('DTM Analysis')
root.geometry('425x620')

def getid():
    global out, filepath
    
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    p1.config(text = 'Path:  ' + str(filenames[0]))
    
    filepath = '\\'.join(filenames[0].split('/')[:-1])
    out = filepath + '\\' + 'Seal Data - Compiled.xlsx'
    df = pd.DataFrame()
    print('start')
    
    file = pd.ExcelFile(filenames[0])
    sheets = file.sheet_names
    
    df = pd.concat(pd.read_excel(filenames[0], sheet_name=None), ignore_index=True)
    df = df[['Substation name','Installation ID','Meter No.']]
    if os.path.exists(out) == False: df.to_excel(out, index = False)
    
    # if both files are already in, prompt to click button
    if os.path.exists(filepath + '\\State Data - Cleaned.csv',): 
        textbox.delete('1.0', END)
        textbox.insert(END, ' Click the \'Clean SEAL Data\' button')
    
def clean():

    states = fd.askopenfilenames()
    p2.config(text = 'Path:  ' + str(states[0]))
    
    filepath = '\\'.join(states[0].split('/')[:-1])
    t = open(filepath + '\\' + 'temp.txt','w', encoding = 'utf-8')
    
    with open(states[0], 'r') as f:
        for line in f.readlines():
            if '|' in line:
                if len(line.replace('|','').replace('-','')) > 5:
                    t.write(line)
                
    f.close()
    t.close()
    print(filepath)
    
    df = pd.read_csv(filepath + '\\temp.txt', sep = '|')
    df = df.drop_duplicates()
    
    # delete rows that mimic header
    df = df.drop(df.index[df['State'] == 'State'].tolist())
    df = df[df.columns[1:-1]]
    df = df.rename(columns = lambda x: x.strip())
    df = df[['Installation','Device No.','Contract Acc.']]
    
    os.remove(filepath + '\\temp.txt')
    df.to_csv(filepath + '\\State Data - Cleaned.csv', index = False)
    
    # if both files are already in, prompt to click button
    if os.path.exists(filepath + '\\Seal Data - Compiled.xlsx'): 
        textbox.delete('1.0', END)
        textbox.insert(END, ' Click the \'Clean SEAL Data\' button')
    
def combine():
    global out
    
    filepath = '\\'.join(out.split('\\')[:-1])
    writer = pd.ExcelWriter(out,  mode = 'a', engine = 'openpyxl')
    print(out)
    print(filepath)
    dfseal = pd.read_excel(out)
    dfstate = pd.read_csv(filepath + '\\State Data - Cleaned.csv')
    
    # merge again using inst
    dfstate = dfstate.rename(columns = {'Installation':'Installation ID'})
    dfout = pd.merge(left = dfseal, right = dfstate, how = 'left', on = 'Installation ID')
    dfout.loc[(pd.isnull(dfout['Device No.'])), 'Device No.'] = dfout['Meter No.']
    dfout = dfout.drop(columns = ['Meter No.'])
    dfout = dfout.loc[(pd.isnull(dfout['Installation ID']) == False)]
    
    if 'Clean' in writer.book.sheetnames: writer.book.remove(writer.book['Clean'])
    
    dfout['Device No.'] = dfout['Device No.'].astype(str).apply(lambda x: x.strip())
    dfout.to_excel(writer, sheet_name = 'Clean', index = False)
    writer.close()
    print('woo')

    # output BCRM and SQL inputs
    output = dfout['Contract Acc.'].drop_duplicates().dropna().astype(str).tolist()
    outp.config(text = 'Find kWh reading on BCRM ZBI and upload')
    act_total.config(text = '')
    
    textvar2.set('BCRM')
    
    textbox.delete('1.0', END)
    for i in output:
        if '.' in i: i = i[:-2]
        textbox.insert(END, i + '\n')
    
    total.config(text = 'Displayed IDs = ' + str(len(output)))

def outid2():

    global out, all_ids, ind

    outdf = pd.read_excel(out, sheet_name = 'Clean')
    ind = 0
    all_ids = []
    textbox.delete('1.0', END)
    
    if textvar2.get() == 'SQL':
        b3["state"] = DISABLED
        outdf = outdf.loc[(outdf['Device No.'].str.len() > 15)]
        new_ids = outdf['Device No.'].drop_duplicates().tolist()
        all_ids.append(new_ids)
        act_total.config(text = 'Total IDs = ' + str(len(new_ids)))

        if len(new_ids) > 1000: divide(new_ids, 1000)

        textbox.insert(END,'\'' + '\',\''.join(all_ids[ind]) + '\'')  
        
        total.config(text = 'Displayed IDs = ' + str(len(all_ids[ind])))
        outp.config(text = 'Find kWh reading on SQL & upload for both dates')
        
    else:
        b3["state"] = DISABLED
        b4["state"] = DISABLED

        new_ids = outdf['Contract Acc.'].drop_duplicates().dropna().astype(str).tolist() 
        act_total.config(text = '')
       
        for i in new_ids:
            if '.' in i: i = i[:-2]
            textbox.insert(END, i + '\n')
        
        total.config(text = 'Displayed IDs = ' + str(len(new_ids)))
        outp.config(text = 'Find kWh reading on BCRM ZBI and upload')

def comb(event):
    outid2()

def getcons():
    global out, all_ids, ind
    
    textbox.delete('1.0', END)
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    p3.config(text = 'Path: ' + str(filenames[0]))
    
    zbi = pd.read_excel(filenames[0])
    
    if cal.get_date() == date.today(): textbox.insert(END, ' Select Start and End Date, and reupload')
    
    else:
        print('son')
        
        noofdays = int((cal2.get_date() - cal.get_date()).days)
        
        zbi['Cust. Reading'] = (zbi['Current usage consumption'].astype(float) / zbi['Bill duration'].astype(int) * noofdays).astype(int)
        zbi = zbi.loc[(zbi['Print Date'].dt.month == cal.get_date().month) & (zbi['Print Date'].dt.year == cal.get_date().year), ['Contract Account','Cust. Reading']]
        zbi = zbi.rename(columns = {'Contract Account':'Contract Acc.'})
        print(zbi)
        
        writer = pd.ExcelWriter(out,  mode = 'a', engine = 'openpyxl')
        outdf = pd.read_excel(out, sheet_name = 'Clean')
        
        outdf = pd.merge(left = outdf, right = zbi, how = 'left', on = 'Contract Acc.')
        outdf = outdf.drop_duplicates()
        
        writer.book.remove(writer.book['Clean'])
        outdf.to_excel(writer, sheet_name = 'Clean', index = False)
        
        writer.close()
        all_ids = []
        ind = 0
        
        # output meter IDs for SQL
        if 'SM Reading' in outdf.columns: textbox.insert(END,' Upload DTM Consumption File')  
        
        else:
            outdf = outdf.loc[(outdf['Device No.'].str.len() > 15)]
            new_ids = outdf['Device No.'].drop_duplicates().astype(str).tolist()
            all_ids.append(new_ids)
            act_total.config(text = 'Total IDs = ' + str(len(new_ids)))
            
            if len(new_ids) > 1000: divide(new_ids, 1000)
            textbox.insert(END,'\'' + '\',\''.join(all_ids[ind]) + '\'')  
            total.config(text = 'Displayed IDs = ' + str(len(all_ids[ind])))
            outp.config(text = 'Find kWh reading on SQL & upload for both dates')
            textvar2.set('SQL')
    
def getsql():
    
    global out
    
    textbox.delete('1.0', END)
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    p4.config(text = 'Path: ' + str(filenames))
    
    writer = pd.ExcelWriter(out,  mode = 'a', engine = 'openpyxl')
    outdf = pd.read_excel(out, sheet_name = 'Clean')
    bigsq = pd.DataFrame()

    for i in filenames:
        sq = pd.read_excel(i)
        print(len(sq))
        
        # reformat
        dates = pd.to_datetime(sq.loc[1,'READ_TIME'].split(' ')[0], format = '%d-%b-%y').strftime('%d/%m/%Y')
        sq = sq[['METER_ID','READ_VALUE']]
        sq = sq.rename(columns = {'METER_ID':'Device No.','READ_VALUE':dates})
        # biqsq = pd.concat([bigsq,sq]) if len(bigsq) > 5 else sq
        outdf = pd.merge(left = outdf, right = sq, how = 'left', on = 'Device No.')
        
        for j in outdf.columns:
            if '_x' in j: 
                outdf.loc[(pd.isnull(outdf[j])), j] = outdf[j.split('_')[0] + '_y']
                outdf = outdf.rename(columns = {j:j.split('_')[0]})
                outdf = outdf.drop(columns = [j.split('_')[0] + '_y'])
        
    # print(bigsq)
    # outdf = pd.merge(left = outdf, right = bigsq, how = 'left', on = 'Serial Number')
    if cal.get_date() == date.today(): textbox.insert(END, ' Select Start and End Date, and reupload')
    
    elif cal.get_date().strftime('%d/%m/%Y') in outdf.columns and cal2.get_date().strftime('%d/%m/%Y') in outdf.columns:
        
        # is it okay if null values are put as 0? should be ??
        # outdf[cal.get_date().strftime('%d/%m/%Y')] = outdf[cal.get_date().strftime('%d/%m/%Y')].apply(lambda x: int(x) if x)
        # outdf[cal2.get_date().strftime('%d/%m/%Y')] = outdf[cal2.get_date().strftime('%d/%m/%Y')].apply(lambda x: int(x) if x)
        
        tempdf = outdf.loc[(pd.isnull(outdf[cal2.get_date().strftime('%d/%m/%Y')]) == False) & (pd.isnull(outdf[cal.get_date().strftime('%d/%m/%Y')]) == False)]
        
    
        outdf['SM Reading'] = (tempdf[cal2.get_date().strftime('%d/%m/%Y')].astype(int) - tempdf[cal.get_date().strftime('%d/%m/%Y')]).astype(int)
        # outdf.loc[(pd.isnull(outdf['SQL Diff.']) == False), 'SM Reading'] = outdf['SM Reading'].astype(int)
        
        if 'Cust. Reading' in outdf.columns: textbox.insert(END,' Upload DTM Consumption File') 
                
        else: 
            output = outdf['Contract Acc.'].drop_duplicates().dropna().astype(str).tolist()
            for i in output:
                if '.' in i: i = i[:-2]
                textbox.insert(END, i + '\n')
                
        writer.book.remove(writer.book['Clean'])
        outdf = outdf.drop_duplicates()
        outdf.to_excel(writer, sheet_name = 'Clean', index = False)

        writer.close() 

def getdtm():
    global dtmpath
    
    textbox.delete('1.0', END)
    dtmpath = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    p5.config(text = 'Path: ' + str(dtmpath[0]))

    dtmpath = dtmpath[0]
    
    # bigtemp = []

    # for i in dtmpath:
    #     pe = i.split('.')[0].split('/')[-1]
    #     df = pd.read_excel(i)
        
    #     # date column
    #     df['date_val'] = pd.to_datetime(df['date_val'], format = '%m%d%y')
    #     df = df.sort_values(['date_val'], ascending = [True])

    #     dates = df['date_val'].drop_duplicates()

    #     for j in dates:
    #         summ = 0
    #         temp = []
            
    #         temp.append(pe)
    #         temp.append(j)
            
    #         for k in range(1,4): summ += df.loc[(df['date_val'] == j), 'channel_' + str(k) + '_kwh'].sum()
    #         temp.append(summ)
    #         bigtemp.append(temp)

    # dtmdf = pd.DataFrame(bigtemp, columns = ['PE','Date', 'Total Sum kWh'])  
    # dtmdf.to_excel('DTM Cons.xlsx', index = False)
    # print(dtmdf)
    
    if cal.get_date() == date.today(): textbox.insert(END, ' Select Start and End Date, and reupload')
    else: textbox.insert(END,' Click button to generate report!')

def divide(array,lim):

    global all_ids, ind

    all_ids = []
    for i in range(math.ceil(len(array)/lim)):
        temp = array[lim*i:lim*(i+1)]
        all_ids.append(temp)
    
    b4["state"] = NORMAL   
    return all_ids
    
def nextt():

    global all_ids, ind
       
    b3["state"] = NORMAL
    b4["state"] = DISABLED
    ind += 1
    
    textbox.delete('1.0', END)
    
    if textvar2.get() == 'BCRM':
        for i in all_ids[ind]:
            if '.' in i: i = i[:-2]
            textbox.insert(END, i + '\n')
        outp.config(text = all_ids[ind])
            
    else: textbox.insert(END,'\'' + '\',\''.join(all_ids[ind]) + '\'')  
    
    if ind < len(all_ids)-1:
        b4["state"] = NORMAL

    total.config(text = 'Displayed IDs = ' + str(len(all_ids[ind])))
   
def back():
    global all_ids, ind
    
    b4["state"] = NORMAL
    b3["state"] = DISABLED
    ind -= 1
    
    textbox.delete('1.0', END)
    
    if textvar2.get() == 'BCRM':
        for i in all_ids[ind]:
            if '.' in i: i = i[:-2]
            textbox.insert(END, i + '\n')
        outp.config(text = all_ids[ind])
        
    else: textbox.insert(END,'\'' + '\',\''.join(all_ids[ind]) + '\'')
    
    if ind != 0:
       b3["state"] = NORMAL 
   
    total.config(text = 'Displayed IDs = ' + str(len(all_ids[ind])))

def genrep():
    global out, dtmpath
    
    df = pd.read_excel(out, sheet_name = 'Clean')
    df['Substation name'] = df['Substation name'].apply(lambda x: x.replace('/',''))
    
    filepath = '\\'.join(out.split('\\')[:-1])
    print(filepath)
    print('----------')
    
    wb = openpyxl.Workbook()
    wb.save(filepath + '\\DTM Analysis.xlsx')
    
    writer = pd.ExcelWriter(filepath + '\\DTM Analysis.xlsx', engine = 'xlsxwriter')
    summary = []
    
    # pes = df['Substation Name'].drop_duplicates().tolist()
    
    # read dtm file
    dftm = pd.read_excel(dtmpath, sheet_name = 'Total Sum kWh')
    dftm['Date'] = pd.to_datetime(dftm['Date'], format = '%Y-%m-%d')
    dtm = dftm.loc[(dftm['Date'].dt.date >= cal.get_date()) & (dftm['Date'].dt.date <= cal2.get_date())]
    dtm = dtm.rename(columns = lambda x: x.strip())
    dtm['Source.Name'] = dtm['Source.Name'].str.slice(stop = -4)
    
    dtm.to_excel(writer, sheet_name = 'DTM Data', index = False)
    
    
    print(dtm)
    summary = []
    
    for i in df['Substation name'].drop_duplicates().tolist():
        pedf = df.loc[(df['Substation name'] == i)]
        pedf['kWh Reading'] = pedf['SM Reading']
        pedf.loc[(pd.isnull(pedf['kWh Reading'])) & (pd.isnull(pedf['Cust. Reading']) == False), 'kWh Reading'] = pedf['Cust. Reading']
        # pedf = pedf.rename(columns = {'SM Reading':'kWh Reading'})
        # pedf = pedf[['Installation ID','Device No.','Contract Acc.','kWh Reading']]
        
        print(i)
        temp = dtm.loc[(dtm['Source.Name'].str.lower() == i.lower())]
            # summary.append([i, temp['Total Sum kWh'].sum()]) # add dtm data too
        print(temp)
        parent = temp['Total Sum kWh'].sum()
        print(parent)
        child = pedf['kWh Reading'].sum()
        diff = parent-child
        
        div = round(diff/parent*100,2) if parent != 0 else None
        
        
        summary.append([i, parent, child, diff, str(div) + '%']) # add dtm data too
        
        pedf = pedf.drop_duplicates()
        pedf.to_excel(writer, sheet_name = i, index = False)

    # summary sheet
        # # make dataframe to take all rows between the two dates
    # else:  
        # 
        
        # for i in dtm['Source.Name'].drop_duplicates().tolist():
            
        
    # # make loop for all PEs
    # # sum up total kwh for each
   
    
    
    summ = pd.DataFrame(summary, columns = ['PE Name','Parent Consumption (kWh)','Child Consumption (kWh)','Diff kWh', '% Diff'])
    summ.loc[(summ['Diff kWh'] < 0), ['Diff kWh', '% Diff']] = None
    if 'Summary' in writer.book.sheetnames: writer.book.remove(writer.book['Summary'])
    summ.to_excel(writer, sheet_name = 'Summary',index = False)
    
    writer.close()
    
    textbox.delete('1.0', END)
    textbox.insert(END, ' DTM Analysis report has been created.')
    
    b4["state"] = DISABLED
    b3["state"] = DISABLED
    
    outp.config(text = '')
    total.config(text = '')
    act_total.config(text = '')
    
def exitt(event):
    root.quit()
    

Button(root, text = 'SEAL Data', command = getid).place(x = 35, y = 70) 
Button(root, text = 'State Data', command = clean).place(x = 34, y = 110) 
Button(root, text = 'Get Cons. (kWh)', command = getcons).place(x = 28, y = 150) 
Button(root, text = 'Get SQL Reading', command = getsql).place(x = 26, y = 190) 
Button(root, text = 'DTM Consumption', command = getdtm).place(x = 26, y = 230) 
Button(root, text = 'Clean SEAL Data', command = combine).place(x = 95, y = 275) #report
Button(root, text = 'Generate DTM Analysis', command = genrep).place(x = 195, y = 275) #report

p1 = Label(root, text = 'Path: ')
p2 = Label(root, text = 'Path: ')
p3 = Label(root, text = 'Path: ')
p4 = Label(root, text = 'Path: ')
p5 = Label(root, text = 'Path: ')

Label(root, text = 'Start Date: ').place(x = 75, y = 30)
Label(root, text = 'End Date: ').place(x = 210, y = 30)

# cal = Calendar(root)
cal = DateEntry(root, width = 7)
cal.place(x = 135, y = 30)

cal2 = DateEntry(root, width = 7)
cal2.place(x = 270, y = 30)

b3 = Button(root, text = 'Back', command = back)
b4 = Button(root, text = 'Next', command = nextt)

p1.place(x = 110, y  = 72)
p2.place(x = 110, y  = 112)
p3.place(x = 140, y  = 152)
p4.place(x = 140, y  = 192)
p5.place(x = 150, y  = 232)
b3.place(x = 30, y = 577)
b4.place(x = 67, y = 577)

textvar2 = StringVar()
choices = ['SQL','BCRM']

cb = ttk.Combobox(root, textvariable = textvar2, values = choices, width = 6)
cb.bind('<<ComboboxSelected>>', comb)
cb.place(x = 315, y = 320)

outp = Label(root, text = 'Output')
total = Label(root, text = '')
act_total = Label(root, text = '')

outp.place(x = 30, y = 320)
total.place(x = 265, y = 580)
act_total.place(x = 105, y = 580)

b3["state"] = DISABLED
b4["state"] = DISABLED

textbox = ScrolledText(root, height = 14, width = 43)
textbox.place(x = 30, y = 346)

textbox.insert(END, ' Upload files for SEAL and State Data')
Label(root, text = 'SAB').place(x = 395, y = 600)

root.resizable(True, False) 
root.bind('<Escape>', exitt)
root.mainloop()